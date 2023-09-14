import os
import openpyxl
from openpyxl import load_workbook
import datetime

fileToParseList = []
filesProcessed = 0
filesFound = 0
for dirpath, dirs, files in os.walk(r'C:\Work\Dev\P472 Production Livelink'):
  for filename in files:
    fname = os.path.join(dirpath,filename)
    if '_M.xlsx' in fname:
        filesProcessed += 1
        try:
            wb = load_workbook(filename=fname)
            try:
                sheetList = wb.sheetnames
                if 'ALL Machines' in sheetList:
                    fileToParseList.append(fname)
                    filesFound += 1
            except:
                pass
            finally:
                wb.close()
        except:
            print(f'Cannot open file {fname}')
    print(f"Processed {filesProcessed}, Found {filesFound}")

print(fileToParseList)
currentTime = datetime.datetime.now()
try:
    filesProcessed = open('files_processed.txt', 'w')
    filesProcessed.write(f"Last Access [{currentTime}]")
    filesProcessed.writelines(fileToParseList)
except Exception as e:
    print(f"Error maintaining file list {e}")
finally:
    try:
        filesProcessed.close()
    except:
        print("File already closed")
